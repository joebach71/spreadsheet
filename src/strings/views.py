# Create your views here.
# from .models import *
from .models import PRODUCTS, Product
from django.shortcuts import render
from django.http.response import HttpResponseRedirect, HttpResponseForbidden,\
    JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.utils import timezone
from django.db.models import F, Q
from django.http import Http404, HttpResponseServerError, HttpResponseBadRequest

# export json
import json
from deepdiff import DeepDiff
import re, logging, os#, io
from django.core.serializers import serialize
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view, renderer_classes
from rest_framework.response import Response
from rest_framework.renderers import JSONRenderer
from rest_framework.reverse import reverse
from rest_framework import status, permissions, generics, serializers, viewsets
from rest_framework.pagination import PageNumberPagination

from strings.serializers import ProductSerializer, UserSerializer, GroupSerializer, SERIALIZERS, LOCALSERIALIZERS
from strings.permissions import IsUserAdminOrReadOnly
from django.contrib.auth.models import Group

from languagestrings.settings import MEDIA_ROOT, APP_NAME

from .parser import dump2dict, get_dict_item_changed, get_dict_item_value_changed, formatColumns
from .validator import validateHTMLFragment
'''
export CDN
'''
#from subprocess import Popen, PIPE
# paginate    
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django_filters.rest_framework import DjangoFilterBackend, FilterSet, NumberFilter
from rest_framework import filters

# export as excel from data
import openpyxl
import datetime
logger = logging.getLogger('strings')

login_url = '/login/'

VIEWS = {}
LOCALVIEWS = {}

class StandardResultsSetPagination(PageNumberPagination):
    page_size = None
    page_size_query_param = 'page_size'

def fabricViewSet(myClass, serializeClass):
    class myClassViewSet(viewsets.ModelViewSet):
        queryset = myClass.objects.all()
        serializer_class = serializeClass
        pagination_class = StandardResultsSetPagination
        ordering_fields = ('__all__')
        ordering = ('stringid','English')
        filter_backends = (DjangoFilterBackend,filters.OrderingFilter,)
        filter_fields = {
            'stringid': ['icontains',],
            'description1': ['icontains',],
            'description2': ['icontains',],
            'Korea': ['icontains',],
            'English': ['icontains',],
            'China': ['icontains',],
            'Indonesia': ['icontains',],
            'Thailand': ['icontains',],
            'Vietnam': ['icontains',],
        }
        def get_queryset(self):
            """
            Optionally restricts the returned purchases to a given user,
            by filtering against a `username` query parameter in the URL.
            """
            queryset = self.queryset
            untranslated = self.request.query_params.get('untranslated', False)
            unmodified = self.request.query_params.get('unmodified', False)
            user = User.objects.get(username=self.request.user.username)
            if user is not None:
                groups = user.groups.all()
                qs_params = None
                if untranslated:
                    for group in groups:
                        if group.name != 'admin' and group.name !='manager':
                            q = Q(**{ group.name+'__exact' : "" })
                            if qs_params:
                                qs_params = qs_params | q
                            else:
                                qs_params = q
                elif unmodified:
                    qs_exclude = None
                    for group in groups:
                        if group.name != 'Korea' and group.name != 'admin' and group.name !='manager':
                            qx = Q(**{ group.name+'__exact' : ""})
                            if qs_exclude:
                                qs_exclude = qs_exclude & qx
                            else:
                                qs_exclude = qx
                            q = Q(**{ group.name+'_modified_at__lte' : F('Korea_modified_at') })
                            if qs_params:
                                qs_params = qs_params | q
                            else:
                                qs_params = q
                    if qs_exclude is not None:
                        queryset = queryset.exclude(qs_exclude)
                if qs_params is not None:
                    queryset = queryset.filter(qs_params)
            return queryset
    return myClassViewSet

for product, myClass in PRODUCTS.iteritems():
    VIEWS[product] = fabricViewSet(myClass, SERIALIZERS[product])

def save_uploaded_file(f, filename):
    with open(filename, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    destination.close()
    return

@login_required
def import_data(request, pk):
    if not request.user.is_authenticated():
        return HttpResponseRedirect("/login/")

    user = User.objects.get(username=request.user.username)
    try:
        document = Product.objects.get(pk=pk)
    except:
        return Http404
   
    # not valid request
    if request.method != 'POST':
        return Http404
    if not request.FILES['fileToUpload']:
        return Http404
    if document.category != request.POST['category']:
        return Http404
    timestamp = timezone.now()
    filename = os.path.join(MEDIA_ROOT, '{0}_{1}_{2}.xlsx'.format(document.category, \
        user.username, timestamp.strftime('%Y%m%d%H%M%S')))
    logger.info("Uploaded '{0}' file saved as {1}".format(request.FILES['fileToUpload'].name, filename))
    save_uploaded_file(request.FILES['fileToUpload'], filename)
    # convert excel file into json
    fields = []
    for group in user.groups.all():
        if group.name == 'admin' or group.name == 'manager':
            continue
        fields.append(group.name)
    columns = formatColumns(fields)
    importData = dump2dict(filename, document.category, columns) # data from excel import
    serverData = get_dynamic_dict_data(document, None, None, fields) # data from table
    # logger.debug('import %s' % importData.get('update'))
    # logger.debug('server %s' % serverData)
    # compare the diff
    if not importData:
        # Format is wrong prob.
        c = {
            'columns': columns,
            'name': request.FILES['fileToUpload'].name,
            'error_message': 'Unable to parse Excel file'
        }
        return render(request, 'strings/error_import.html', c)

    diff = DeepDiff(serverData, importData['update'])
    # logger.debug('diff %s' % diff)
    changes = {
        'error': False,
        'inValid': importData.get('in_valid'),
        'empty': importData.get('empty'),
        'add': get_dict_item_changed(diff.get('dictionary_item_added'), importData.get('update')) or [],
        'duplicate': importData.get('dup'),
        'update': get_dict_item_value_changed(diff.get('values_changed')),
        'delete': get_dict_item_changed(diff.get('dictionary_item_removed'), serverData),
    }
    if len(changes.get('inValid')) or len(changes.get('duplicate')) or len(changes.get('empty')) or \
        len(changes.get('add')):
        changes['error'] = True
    logger.debug('Changes: %s' % changes)

    c = {
        'columns': columns,
        'name': request.FILES['fileToUpload'].name,
        'data': changes,
        'table': document,
        # 'SiteHeader': 'Translations Management',
        'page_title': 'Import Confirm Page',
    }
    return render(request, 'strings/import.html', c)
import_data.short_description = u"Import XLSX Data"

@login_required
def confirm_import_data(request, pk):
    if not request.user.is_authenticated():
        return HttpResponseRedirect("/login/")

    user = User.objects.get(username=request.user.username)
    groups = user.groups.all()
    try:
        document = Product.objects.get(pk=pk)
    except:
        return Http404
   
    # not valid request
    if request.method != 'POST':
        return Http404

    if not request.POST:
        return HttpResponseServerError('Form Not Found')
    
    timestamp = timezone.now()
    for name in request.POST:
        # logger.debug(name)
        if name == 'csrfmiddlewaretoken':
            continue
        value = request.POST.get(name, None)
        # split into stringid and field
        column = name.split('.', 2)
        try:
            item = PRODUCTS["{0}_{1}".format(document.name, document.category)].objects.get(pk=column[0])
        except (KeyError, PRODUCTS["{0}_{1}".format(document.name, document.category)].DoesNotExist) as e:
            logger.error(e)
            return HttpResponseServerError(e)
        existing_value = getattr(item, column[1])
        if value == existing_value:
            continue
        # modifiedAt = "{0}_modified_at".format(column[1])
        setattr(item, column[1], value)
        # setattr(item, modifiedAt, timestamp)
        item.save()
    logger.info(user.get_username()+' successfully import at '+str(timestamp))
    return HttpResponseRedirect("/")
confirm_import_data.short_description = u"Confirm Imported XLSX Data"

@login_required
def export_xlsx(request, pk):
    timestamp = timezone.now()
    if not request.user.is_authenticated():
        return HttpResponseRedirect("/login/")
    try:
        document = Product.objects.get(pk=pk)
    except:
        return Http404

    wb = openpyxl.Workbook()
    # ws = wb.get_active_sheet()
    ws = wb.active
    ws.title = "{0}".format(document.category)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={0}.xlsx'.format(ws.title)

    columns = [u'stringid',u'Korea',u'English']
    user = User.objects.get(username=request.user.username)
    logger.info(user.get_username()+' requested export '+document.category+' at '+str(timestamp)) 
    groups = user.groups.all()
    for group in groups:
        if group.name != 'Korea' and group.name != 'English' and group.name != 'admin' and group.name != 'manager':
            columns.append(group.name)
    row_num = 1
    
    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num, column=col_num + 1)
        if columns[col_num] == 'stringid':
            c.value = 'id'
        else:
            # lower case and first 3 letters
            c.value = columns[col_num][0:3].lower()
    
    # untranslated only
    untranslated = True if request.GET.get('untranslated') == 'true' else False
    key = request.GET.get('filter')
    pattern = request.GET.get('pattern')
    # logger.info('Raw Data: "%s"' % request.body)
    # logger.info('fields: %s' % columns)
    # get queryset from serializer
    if (untranslated == True):
        queryset = get_dynamic_queryset(PRODUCTS["{0}_{1}".format(document.name, document.category)], key, pattern, columns)
    else:
        queryset = get_dynamic_queryset(PRODUCTS["{0}_{1}".format(document.name, document.category)], key, pattern)

    for obj in queryset:
        row_num += 1
        row = []
        for attr in columns:
            row.append(getattr(obj, attr))
        
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num, column=col_num + 1)
            c.value = row[col_num]
    wb.save(response)
    return response
export_xlsx.short_description = u"Export XLSX"

@login_required
def confirm(request, pk):
    if not request.user.is_authenticated():
        return HttpResponseRedirect(reverse('languagestrings:login'))
    if not request.user.is_active:
        return HttpResponseForbidden
    
    try:
        document = Product.objects.get(pk=pk)
    except:
        return Http404
    

    '''
    Process all parameters
    '''
    tablename = "{}_{}".format(document.name, document.category)
    # Always return an HttpResponseRedirect after successfully dealing
    # with POST data. This prevents data from being posted twice if a
    # user hits the Back button.
    saved = []
    reset = []
    notSaved = []
    notFound = []
    logger.info('Confirmed Raw data: "%s"' % request.body.decode("utf-8"))
    data = json.loads(request.body.decode('utf-8'))
    for item in data:
        stringid = item[0]
        attribute = item[1]
        before = item[2]
        after = item[3]
        try:
            languagestring = PRODUCTS[tablename].objects.get(pk=stringid)
            stored = getattr(languagestring, attribute)
            if type(stored) == datetime.datetime:
                after = timezone.now();
                item[3] = '{}'.format(after)
                try:
                    setattr(languagestring, attribute, after)
                    languagestring.save()
                    # confirmation
                    reset.append(item)
                except Exception as e:
                    logger.error('Unabled to Save "{}", "{}"'.format(item,e))
                    item.append(str(e))
                    notSaved.append(item)
            elif (stored == before):
                setattr(languagestring, attribute, after)
                try:
                    languagestring.save()
                    # confirmation
                    saved.append(item)
                except Exception as e:
                    logger.error('Unabled to Save "{}", "{}"'.format(item,e))
                    item.append(str(e))
                    notSaved.append(item)
            else:
                # field value has been changed... must be verified
                logger.error('Database value has changed from previous. expected:{}::stored:{}'.format(before, stored))
                item.append('Current Server Data is different from before data - {}'.format(stored))
                notSaved.append(item)
        except (KeyError, PRODUCTS[tablename].DoesNotExist) as e:
            logger.error('Unable to retrieve record:{}'.format(e))
            item.append(e)
            notFound.append(item)
    if (len(notSaved)):
        returncode = "500"
    else:
        '''
            Commit Changes to version control
        '''
        returncode = "200"
    # svnMgr = document.commit_to_vc(request.user.username)
    #print svnMgr.diff()
    context = {
        'saved': saved,
        'reset': reset,
        'notSaved': notSaved,
        'notFound': notFound,
        'returncode': returncode,
        'product': pk,
    }
    return HttpResponse(json.dumps(context), content_type="application/json")

@login_required
def save(request, pk):
    if not request.user.is_authenticated():
        return HttpResponseRedirect(reverse('languagestrings:login'))
    if not request.user.is_active:
        return HttpResponseForbidden
    
    try:
        user = User.objects.get(username=request.user.username)
    except (KeyError, User.DoesNotExist):
        '''
        Should this provide error message?
        '''
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    '''
    Process all parameters
    '''
    # Always return an HttpResponseRedirect after successfully dealing
    # with POST data. This prevents data from being posted twice if a
    # user hits the Back button.
    
    try:
        document = Product.objects.get(pk=pk)
    except:
        return Http404
    
    reset = []
    result = []
    errors = []
    abnormal = {}
    notFound = {}
    if not request.is_ajax():
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

    logger.info('Raw data: "%s"' % request.body.decode("utf-8"))
    data = json.loads(request.body.decode("utf-8"))
    for stringidField, value in data.iteritems():
        stringid = value['stringid']
        field = value['field']

        languagestring = PRODUCTS["{0}_{1}".format(document.name, document.category)].objects.get(pk=stringid)
        translation = getattr(languagestring, field)
        if type(translation) == datetime.datetime:
            value['after'] = '{}'.format(timezone.now())
            reset.append(value)
        else:
            if (translation != value['before']):
                # field value has been changed... must be verified
                abnormal[stringid] = value
                abnormal[stringid]['server'] = translation
                continue
            if (translation == value['after']):
                continue
            # validate html
            value['parsed'], err = validateHTMLFragment(value['after'])
            if (err):
                logger.error('[ERROR] HTML Validation: "%s"' % err)
                value['reason'] = err;
                errors.append(value)
            else:
                result.append(value)        
        logger.info('Response: "%s"' % value);
    context = {
        'data': result,
        'reset': reset,
        'error': errors,
        'product': pk,
    }
    return HttpResponse(json.dumps(context), content_type="application/json") 

def get_dynamic_queryset(myClass, key=None, pattern=None, fields=[]):
    queryset = myClass.objects.all().order_by('stringid')
    if len(fields):
        qs_params = None
        for field in fields:
            # logger.debug(group.name)
            q = Q(**{ field+'__exact' : "" }) | Q(**{ field+'__exact': None})
            if qs_params:
                qs_params = qs_params | q
            else:
                qs_params = q
        queryset = queryset.filter(qs_params)
    if key is not None and pattern is not None:
        kwargs = { '{0}__icontains'.format(key): pattern }
        queryset = queryset.filter(**kwargs)
    return queryset

def get_dynamic_dict_data(product, key=None, pattern=None, fields=[]):
    queryset = PRODUCTS["{0}_{1}".format(product.name, product.category)].objects.all().order_by('stringid')
    records = {}
    for obj in queryset:
        data = {}
        data['key'] = obj.pk # should the key part of object - change you must change import data as well 
        for field in fields:
            if getattr(obj, field) is None:
                data[field] = u""
            else:
                data[field] = getattr(obj, field)
        records[obj.pk] = data
    return records

@login_required
def index(req):
    if not req.user.is_authenticated():
        return HttpResponseRedirect("/login/")
    if not req.user.is_active:
        return HttpResponseForbidden
    template_name = 'strings/index.html'
    user = User.objects.get_by_natural_key(req.user.username)
    products = Product.objects.filter(status="p")

    publishers = []
    editDesc = editKorea = editEnglish = admin = False
    for group in user.groups.all():
        if group.name == 'admin':
            admin = True
        elif group.name == 'English':
            editEnglish = True
        elif group.name == 'Korea':
            editKorea = True
        elif group.name == 'manager':
            # skip this role - only for post synchronization
            pass
        else:
            publishers.append(group.name)
    c = {
        'app_name': APP_NAME,
        'publishers': publishers,
        'editDesc' : editDesc,
        'editKorea' : editKorea,
        'editEnglish' : editEnglish,
        'products': products,
        'page_size': 250,
        'SiteHeader': 'Translations Management WebTool(TMW)',
        'page_title': 'LocalText Translations',
    }
    
    return render(req, template_name, c)

@login_required
def get_languagestring(request, pk, region, stringid):
    if not request.user.is_authenticated():
        return HttpResponseRedirect("/login/")
    context = {}
    try:
        document = Product.objects.get(pk=pk)
    except (KeyError, Product.DoesNotExist):
        return Http404
    try:
        languagestring = PRODUCTS["{0}_{1}".format(document.name, document.category)].objects.get(pk=stringid)
    except (KeyError, PRODUCTS["{0}_{1}".format(document.name, document.category)].DoesNotExist):
        return HttpResponse(json.dump(context), status=404, content_type="application/json")
    context[stringid] = getattr(languagestring, region)
    return HttpResponse(json.dump(context), status=200, content_type="application/json")

def get_context_queryset(model, key=None, pattern=None, groups=[]):
    languagestrings = model.objects.all().order_by('stringid')
    if (len(groups)):
        qs_params = None
        for group in groups:
            if group.name != 'admin':
                q = Q(**{group.name+'__exact': None})
                qs_params = qs_params | q if qs_params else q
        languagestrings = languagestrings.filter(qs_params).order_by('stringid')
    if (key is not None and pattern is not None):
        kwargs = { '{0}__{1}'.format(key, 'icontains') : pattern }
        languagestrings = languagestrings.filter(**kwargs).order_by('stringid')
    return languagestrings

class UserViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = User.objects.all().order_by('-date_joined')
    serializer_class = UserSerializer
    pagination_class = StandardResultsSetPagination

class GroupViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    pagination_class = StandardResultsSetPagination

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    pagination_class = StandardResultsSetPagination
