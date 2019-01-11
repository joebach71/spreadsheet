import logging
import json, re

from tidylib import tidy_fragment
from openpyxl import load_workbook
from .validator import validate_texts
# from .constants import NO_VALID

logger = logging.getLogger('parser')
regex_dict_item_changed = re.compile('root\[\'(.*?)\'\]')
regex_dict_item_value_changed = re.compile('root\[\'(.+?)\'\]\[\'(.+?)\'\]')

def get_worksheet(filename, sheetname):
    # parse xlsx file into dataset
    workbook = load_workbook(filename)
    # for wsname in workbook.get_sheet_names():
    for wsname in workbook.sheetnames:
        if wsname.upper() == sheetname.upper():
            worksheet = workbook[wsname]
            logger.debug('Located worksheet, {0}'.format(worksheet.title))
            break
    return worksheet

def formatColumns(fields):
    columns = []
    for field in fields:
        # print group.name
        name = field.lower()
        '''
        import xlsx column name is not same as datatable column name
        allow mapping - TODO: create config for this mapping
        '''
        columns.append({ 'name': name[0:3], 'fieldname': field })
    return columns

def get_headers(worksheet, allowed_headers):
    headers2columns = {}
    columns = {}
    for i in range(1, worksheet.max_column+1):
        cell = worksheet.cell(row=1, column=i)
        if cell.value is None:
            continue
        for header in allowed_headers:
            if header['name'].upper() == cell.value.upper():
                # logger.info(cell.value)
                # logger.info(cell.column)
                columns[header['fieldname']] = cell.column
            if cell.value.upper() == 'ID':
                # this is key
                headers2columns['key'] = cell.column
    headers2columns['col'] = columns
    # logger.debug('Headers: %s ' % headers2columns)
    return headers2columns
def dump2dict(filename, sheetname, fields):
    '''
    Convert xlsx file into json data
    '''
    worksheet = get_worksheet(filename, sheetname)
    if not worksheet:
        return
    # read column header

    # logger.debug("Allowed fields: {}".format(columns))
    # logger.debug("Key ID Column: {}".format(key_column))
    # allowed_headers = allowedColumns(groups)
    headers = get_headers(worksheet, fields)
    # key_column = 'A'
    if len(headers['col']) < 1:
        logger.error('Match column not found!')
        return None
    # filter only available columns
    '''
    {
        id: {
            Id: <id>
            Korea: <korea>
            China: <china>
            ...
        },
        id2: {
            Id: <id2>,
            Korea: <korea>,
            ...
        }
    }
    '''
    duplicated = []
    in_valid = []
    empty = []
    update = {}
    for i in range(2, worksheet.max_row+1):
        key_text = unicode(worksheet['{}{}'.format(headers['key'], i)].value).strip().upper()
        if key_text is None:
            continue
        item = {'key': key_text }
        for header, value in headers['col'].items():
            if worksheet["{0}{1}".format(value, i)].value is not None:
                item[header] = worksheet["{0}{1}".format(value, i)].value
        if len(item.keys()) < 1:
            empty.append(item)
            logger.error('empty entry: %s' % item)
            continue
        if key_text in update:
            logger.error('duplicated entry: %s' % item)
            duplicated.append(item)
            continue
        result = validate_texts(item)
        # logger.debug("failed: {}".format(result["fail"]))
        if len(result['fail']) > 0:
            logger.error('failed on validation: %s' % item)
            in_valid.append({
                'key': key_text,
                'fail': result["fail"],
                'success': result["success"]
            })
            continue
        # logger.info("success: %s" % item)
        update[key_text] = item
    return {'dup': duplicated, 'in_valid': in_valid, 'empty': empty, 'update': update}

def dump2json(filename, sheetname, allowed_headers):
    '''
    Convert xlsx file into json data
    '''
    data = dump2dict(filename, sheetname, allowed_headers)
    return json.dumps(data, ensure_ascii=True, sort_keys=True)

def get_dict_item_value_changed(diff):
    '''
    DOC: return item value changes only - diff should be only 'values_changed', assumes dictionary
    '''
    if diff is None:
        return None
    changed = []
    for key in sorted(diff):
        # root['<id>']['<field>']
        match = regex_dict_item_value_changed.match(key)
        if (match):
            stringid = match.group(1)
            field = match.group(2)
            changed.append({ 'key': stringid, 'language': field,
                           'new_value': diff[key]['new_value'], 'old_value': diff[key]['old_value'],})
        else:
            # unexpected format
            logger.error('Not Found %s' % key)
            return None
    return changed

def get_dict_item_changed(diff, reference):
    '''
    DOC: return changed items only - diff should be only 'dictionary_item_added' or 'dictionary_item_removed'
    '''
    if diff is None:
        return None

    changed = []
    for key in sorted(diff):
        # root['<id>']
        match = regex_dict_item_changed.match(key)
        if (match):
            stringid = match.group(1)
            changed.append(reference[stringid])
        else:
            # unexpected format
            logger.error('Not Found %s' % key)
    return changed

